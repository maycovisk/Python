import openpyxl

#Criar uma planilha(book)
book = openpyxl.Workbook()
#Como visualizar pagina existentes
print(book.sheetnames)
#Como criar uma pagina
book.create_sheet('Teste')
#Como selecionar uma pagina
teste_page = book['Teste']
teste_page.append(['Equipamento','Memoria','SO'])
teste_page.append(['note','8','windows'])
teste_page.append(['note','4','windows'])
#Salvar planilha
book.save('Planilha de controle.xlsx')

