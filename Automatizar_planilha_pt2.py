import openpyxl

#Carregando arquivo
book = openpyxl.load_workbook('Planilha de controle.xlsx')
#Selecionando uma pagina
teste_page = book['Teste']
#Imprimindo os dados de cada linha
for rows in teste_page.iter_rows(min_row = 2, max_row = 5):
    #for cell in rows:     #imprime celula a celula um abaixo do outro
        #print(cell.value)
    print(f'{rows[0].value}, {rows[1].value}, {rows[2].value}')
#Alterando os dados
for rows in teste_page.iter_rows(min_row = 2, max_row = 5):
    for cell in rows:
        if cell.value == 'note':
            cell.value = 'Notebook'
#Salvar as alterações
book.save('Planilha de controle v2.xlsx')